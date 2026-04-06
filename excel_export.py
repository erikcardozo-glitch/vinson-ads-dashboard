import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Grownax brand colors
NAVY = "0A1628"
NAVY_MID = "1F4E79"
NAVY_LIGHT = "1A3A5C"
ACCENT_BLUE = "2A6DA8"
TEXT_LIGHT = "C8D6E5"
WHITE = "FFFFFF"
GREEN_ACCENT = "28A745"

HEADER_FILL = PatternFill(start_color=NAVY_MID, end_color=NAVY_MID, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=NAVY_MID)
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
DATA_FONT = Font(name="Calibri", size=10)
ADSET_HEADER_FILL = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
ADSET_HEADER_FONT = Font(name="Calibri", bold=True, color=TEXT_LIGHT, size=10)
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'
INT_FORMAT = '#,##0'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def generate_excel(campaigns, week_start, week_end):
    wb = Workbook()

    _create_instructions_sheet(wb)
    _create_weekly_sheet(wb, campaigns, week_start, week_end)
    _create_adsets_sheet(wb, campaigns, week_start, week_end)
    _create_monthly_sheet(wb, campaigns, week_start)
    _create_tiktok_sheet(wb)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _create_instructions_sheet(wb):
    ws = wb.create_sheet("INSTRUCCIONES", 0)

    ws.merge_cells("A1:D1")
    ws["A1"] = "VINSON — ADS META & TIKTOK"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:D2")
    ws["A2"] = "Reporte generado automáticamente desde Meta Ads API — Grownax"
    ws["A2"].font = SUBTITLE_FONT

    headers = ["CAMPO", "QUÉ CONTIENE", "FRECUENCIA", "FUENTE"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header_style(ws, 3, 4)

    rows = [
        ["Semana", "Fecha de inicio del período", "Semanal / Custom", "API Meta Ads"],
        ["Campaña", "Nombre exacto de la campaña", "Automático", "API Meta Ads"],
        ["Objetivo", "Conversiones / Tráfico / Alcance", "Automático", "API Meta Ads"],
        ["Audiencia", "Tipo de audiencia (intereses / lookalike / retargeting)", "Automático", "API Meta Ads"],
        ["Inversión ($)", "Gasto total en el período", "Automático", "API Meta Ads"],
        ["Impresiones", "Total de impresiones", "Automático", "API Meta Ads"],
        ["Clicks", "Total de clicks", "Automático", "API Meta Ads"],
        ["Compras", "Compras atribuidas", "Automático", "API Meta Ads"],
        ["Ingresos ($)", "Ingresos atribuidos", "Automático", "API Meta Ads"],
        ["Creativos activos", "Nombres de los creativos corriendo", "Automático", "API Meta Ads"],
    ]
    for r_idx, row in enumerate(rows, 4):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18


def _create_weekly_sheet(wb, campaigns, week_start, week_end):
    ws = wb.create_sheet("META ADS SEMANAL")

    ws.merge_cells("A1:M1")
    ws["A1"] = "META ADS — REPORTE SEMANAL"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Período: {week_start} al {week_end}. CTR, ROAS y CPR se calculan con fórmulas."
    ws["A2"].font = SUBTITLE_FONT

    headers = [
        "SEMANA", "CAMPAÑA", "OBJETIVO", "INVERSIÓN ($)",
        "IMPRESIONES", "CLICKS", "CTR %", "COMPRAS", "INGRESOS ($)",
        "ROAS", "CPR ($)", "CREATIVOS ACTIVOS"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header_style(ws, 3, 12)

    calc_cols = [7, 10, 11]

    for r_idx, camp in enumerate(campaigns, 4):
        ws.cell(row=r_idx, column=1, value=week_start).font = DATA_FONT
        ws.cell(row=r_idx, column=2, value=camp["campaign"]).font = DATA_FONT
        ws.cell(row=r_idx, column=3, value=camp["objective"]).font = DATA_FONT

        ws.cell(row=r_idx, column=4, value=camp["spend"]).font = DATA_FONT
        ws.cell(row=r_idx, column=4).number_format = MONEY_FORMAT

        ws.cell(row=r_idx, column=5, value=camp["impressions"]).font = DATA_FONT
        ws.cell(row=r_idx, column=5).number_format = INT_FORMAT

        ws.cell(row=r_idx, column=6, value=camp["clicks"]).font = DATA_FONT
        ws.cell(row=r_idx, column=6).number_format = INT_FORMAT

        ctr_cell = ws.cell(row=r_idx, column=7)
        ctr_cell.value = f'=IF(E{r_idx}=0,"-",F{r_idx}/E{r_idx})'
        ctr_cell.font = DATA_FONT
        ctr_cell.number_format = PCT_FORMAT

        ws.cell(row=r_idx, column=8, value=camp["purchases"]).font = DATA_FONT
        ws.cell(row=r_idx, column=8).number_format = INT_FORMAT

        ws.cell(row=r_idx, column=9, value=camp["revenue"]).font = DATA_FONT
        ws.cell(row=r_idx, column=9).number_format = MONEY_FORMAT

        roas_cell = ws.cell(row=r_idx, column=10)
        roas_cell.value = f'=IF(D{r_idx}=0,"-",I{r_idx}/D{r_idx})'
        roas_cell.font = DATA_FONT
        roas_cell.number_format = '0.00'

        cpr_cell = ws.cell(row=r_idx, column=11)
        cpr_cell.value = f'=IF(H{r_idx}=0,"-",D{r_idx}/H{r_idx})'
        cpr_cell.font = DATA_FONT
        cpr_cell.number_format = MONEY_FORMAT

        ws.cell(row=r_idx, column=12, value=camp["creatives"]).font = DATA_FONT

        for col in range(1, 13):
            cell = ws.cell(row=r_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col in calc_cols:
                cell.fill = GREEN_FILL

    col_letters = ["A","B","C","D","E","F","G","H","I","J","K","L"]
    widths = [12, 40, 15, 14, 14, 10, 10, 10, 14, 8, 12, 30]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = widths[i]


def _create_adsets_sheet(wb, campaigns, week_start, week_end):
    """Hoja con desglose por conjunto de anuncios."""
    ws = wb.create_sheet("CONJUNTOS DE ANUNCIOS")

    ws.merge_cells("A1:K1")
    ws["A1"] = "META ADS — DESGLOSE POR CONJUNTO DE ANUNCIOS"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Período: {week_start} al {week_end}"
    ws["A2"].font = SUBTITLE_FONT

    headers = [
        "CAMPAÑA", "CONJUNTO", "AUDIENCIA", "INVERSIÓN ($)",
        "IMPRESIONES", "CLICKS", "CTR %", "COMPRAS", "INGRESOS ($)",
        "ROAS", "CPR ($)"
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header_style(ws, 3, 11)

    calc_cols = [7, 10, 11]
    r_idx = 4

    for camp in campaigns:
        for a in camp.get("adsets", []):
            ws.cell(row=r_idx, column=1, value=camp["campaign"]).font = DATA_FONT
            ws.cell(row=r_idx, column=2, value=a["adset"]).font = DATA_FONT
            ws.cell(row=r_idx, column=3, value=a["audience"]).font = DATA_FONT

            ws.cell(row=r_idx, column=4, value=a["spend"]).font = DATA_FONT
            ws.cell(row=r_idx, column=4).number_format = MONEY_FORMAT

            ws.cell(row=r_idx, column=5, value=a["impressions"]).font = DATA_FONT
            ws.cell(row=r_idx, column=5).number_format = INT_FORMAT

            ws.cell(row=r_idx, column=6, value=a["clicks"]).font = DATA_FONT
            ws.cell(row=r_idx, column=6).number_format = INT_FORMAT

            ctr_cell = ws.cell(row=r_idx, column=7)
            ctr_cell.value = f'=IF(E{r_idx}=0,"-",F{r_idx}/E{r_idx})'
            ctr_cell.font = DATA_FONT
            ctr_cell.number_format = PCT_FORMAT

            ws.cell(row=r_idx, column=8, value=a["purchases"]).font = DATA_FONT
            ws.cell(row=r_idx, column=8).number_format = INT_FORMAT

            ws.cell(row=r_idx, column=9, value=a["revenue"]).font = DATA_FONT
            ws.cell(row=r_idx, column=9).number_format = MONEY_FORMAT

            roas_cell = ws.cell(row=r_idx, column=10)
            roas_cell.value = f'=IF(D{r_idx}=0,"-",I{r_idx}/D{r_idx})'
            roas_cell.font = DATA_FONT
            roas_cell.number_format = '0.00'

            cpr_cell = ws.cell(row=r_idx, column=11)
            cpr_cell.value = f'=IF(H{r_idx}=0,"-",D{r_idx}/H{r_idx})'
            cpr_cell.font = DATA_FONT
            cpr_cell.number_format = MONEY_FORMAT

            for col in range(1, 12):
                cell = ws.cell(row=r_idx, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col in calc_cols:
                    cell.fill = GREEN_FILL

            r_idx += 1

    col_letters = ["A","B","C","D","E","F","G","H","I","J","K"]
    widths = [35, 35, 15, 14, 14, 10, 10, 10, 14, 8, 12]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = widths[i]


def _create_monthly_sheet(wb, campaigns, week_start):
    ws = wb.create_sheet("META RESUMEN MENSUAL")

    ws.merge_cells("A1:H1")
    ws["A1"] = "META ADS — RESUMEN MENSUAL"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"] = "Resumen automático por mes. Datos agregados de las campañas."
    ws["A2"].font = SUBTITLE_FONT

    headers = ["MES", "AÑO", "INVERSIÓN TOTAL ($)", "COMPRAS TOTALES",
               "INGRESOS TOTALES ($)", "ROAS PROM", "CPR PROM ($)", "NOTAS"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header_style(ws, 3, 8)

    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    total_spend = sum(c["spend"] for c in campaigns)
    total_purchases = sum(c["purchases"] for c in campaigns)
    total_revenue = sum(c["revenue"] for c in campaigns)
    avg_roas = (total_revenue / total_spend) if total_spend > 0 else 0
    avg_cpr = (total_spend / total_purchases) if total_purchases > 0 else 0

    try:
        parts = week_start.split("/")
        if len(parts) == 3:
            data_month = int(parts[1])
            data_year = int(parts[2])
        else:
            data_month = None
            data_year = 2025
    except Exception:
        data_month = None
        data_year = 2025

    for r_idx, month_name in enumerate(months, 4):
        month_num = r_idx - 3
        ws.cell(row=r_idx, column=1, value=month_name).font = DATA_FONT
        ws.cell(row=r_idx, column=2, value=data_year).font = DATA_FONT

        if data_month and month_num == data_month:
            ws.cell(row=r_idx, column=3, value=total_spend).font = DATA_FONT
            ws.cell(row=r_idx, column=3).number_format = MONEY_FORMAT
            ws.cell(row=r_idx, column=4, value=total_purchases).font = DATA_FONT
            ws.cell(row=r_idx, column=5, value=total_revenue).font = DATA_FONT
            ws.cell(row=r_idx, column=5).number_format = MONEY_FORMAT
            ws.cell(row=r_idx, column=6, value=round(avg_roas, 2)).font = DATA_FONT
            ws.cell(row=r_idx, column=7, value=round(avg_cpr, 2)).font = DATA_FONT
            ws.cell(row=r_idx, column=7).number_format = MONEY_FORMAT
            ws.cell(row=r_idx, column=8, value="Datos del período seleccionado").font = DATA_FONT

        for col in range(1, 9):
            ws.cell(row=r_idx, column=col).border = THIN_BORDER
            ws.cell(row=r_idx, column=col).alignment = Alignment(horizontal="center")

    col_letters = ["A","B","C","D","E","F","G","H"]
    widths = [14, 8, 20, 18, 20, 12, 14, 25]
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = widths[i]


def _create_tiktok_sheet(wb):
    ws = wb.create_sheet("TIKTOK ADS SEMANAL")

    ws.merge_cells("A1:K1")
    ws["A1"] = "TIKTOK ADS — REPORTE SEMANAL (PRÓXIMAMENTE)"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:K2")
    ws["A2"] = "Este canal aún no está activo. La estructura está lista para cuando arranque."
    ws["A2"].font = SUBTITLE_FONT

    headers = ["SEMANA", "CAMPAÑA", "OBJETIVO", "FORMATO", "INVERSIÓN ($)",
               "VISTAS", "CLICKS", "COMPRAS / LEADS", "INGRESOS ($)", "ROAS", "CTA DESTINO"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _apply_header_style(ws, 3, 11)
